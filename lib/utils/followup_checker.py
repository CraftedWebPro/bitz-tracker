import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import sys

try:
    from plyer import notification
    PLYER_OK = True
except ImportError:
    PLYER_OK = False

from lib.config import EXCEL_PATH

def check_followups():
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        followup_date_col = None
        followup_sent_col = None
        name_col = None
        phone_col = None

        for idx, h in enumerate(headers, 1):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(idx)
            if h == "Follow Up Date":
                followup_date_col = col_letter
            elif h == "Follow Up Sent?":
                followup_sent_col = col_letter
            elif h == "Business Name":
                name_col = col_letter
            elif h == "Phone No":
                phone_col = col_letter

        if not followup_date_col:
            print("No Follow Up Date column found")
            return []

        today = datetime.now().strftime("%d-%m-%Y")
        due = []

        for row in ws.iter_rows(min_row=2, values_only=False):
            date_val = None
            sent_val = None
            name_val = ""
            phone_val = ""

            for cell in row:
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(cell.column)
                if col_letter == followup_date_col:
                    date_val = str(cell.value or "").strip()
                elif col_letter == followup_sent_col:
                    sent_val = str(cell.value or "").strip().lower()
                elif col_letter == name_col:
                    name_val = str(cell.value or "")
                elif col_letter == phone_col:
                    phone_val = str(cell.value or "")

            if date_val and sent_val == "yes" and date_val <= today:
                due.append({"name": name_val, "phone": phone_val, "date": date_val})

        return due
    except Exception as e:
        print("Error checking follow-ups: " + str(e))
        return []

def send_notification(due_list):
    if not PLYER_OK:
        print("plyer not installed. Install with: pip install plyer")
        print("Due follow-ups: " + str(len(due_list)))
        for d in due_list:
            print("  - " + str(d["name"]) + " (due: " + str(d["date"]) + ")")
        return

    count = len(due_list)
    if count == 0:
        notification.notify(
            title="Biz Tracker - No Follow-ups Due",
            message="You're all caught up! No follow-ups due today.",
            app_name="Biz Tracker",
            timeout=10
        )
    else:
        names = ", ".join(str(d["name"]) for d in due_list[:3])
        if count > 3:
            names += " and " + str(count - 3) + " more"
        notification.notify(
            title="Biz Tracker - " + str(count) + " Follow-up(s) Due!",
            message="Due: " + names,
            app_name="Biz Tracker",
            timeout=15
        )

if __name__ == "__main__":
    due = check_followups()
    if len(sys.argv) > 1 and sys.argv[1] == "--notify":
        send_notification(due)
    else:
        print("Due follow-ups: " + str(len(due)))
        for d in due:
            print("  " + str(d["name"]) + " - due " + str(d["date"]))

