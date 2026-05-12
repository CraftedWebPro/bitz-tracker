import openpyxl
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime

def create_empty_excel(path):
    """Creates a new Excel file with default headers if it doesn't exist."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Tracker"
    
    headers = ["SL No", "Date", "Business Name", "Website", "Problem", "Type", "Location", "Phone No", "Msg Sent?", "Reply Came?", "Follow Up Sent?", "Follow Up Date", "Converted?", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 30
    
    wb.save(path)
    return wb

def ensure_followup_date_column(path):
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if "Follow Up Date" not in headers:
        col_num = len(headers) + 1
        ws.cell(row=1, column=col_num, value="Follow Up Date")
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_num)].width = 14
        wb.save(path)
        return True
    return False

def load_excel(path, data_only=True):
    if not os.path.exists(path):
        create_empty_excel(path)
    ensure_followup_date_column(path)
    return load_workbook(path, data_only=data_only)

def save_excel(wb, path):
    wb.save(path)

def get_all_businesses(ws):
    businesses = []
    headers = [cell.value for cell in ws[1]]
    row_num = 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            row_num += 1
            continue
        biz = dict(zip(headers, row))
        biz['_row'] = row_num
        businesses.append(biz)
        row_num += 1
    return businesses

def add_business_to_excel(wb, ws, path, data):
    row = ws.max_row + 1
    ws[f"A{row}"] = row - 1
    ws[f"B{row}"] = datetime.now().strftime("%d-%m-%Y")
    ws[f"C{row}"] = data.get("Business Name", "")
    ws[f"D{row}"] = data.get("Website", "")
    ws[f"E{row}"] = data.get("Problem", "")
    ws[f"F{row}"] = data.get("Type", "")
    ws[f"G{row}"] = data.get("Location", "")
    ws[f"H{row}"] = data.get("Phone No", "")
    
    col_count = ws.max_column
    if col_count >= 14:
        ws[f"I{row}"] = "No"
        ws[f"J{row}"] = "No"
        ws[f"K{row}"] = "No"
        ws[f"L{row}"] = ""
        ws[f"M{row}"] = "No"
        ws[f"N{row}"] = data.get("Notes", "")
    else:
        ws[f"L{row}"] = "No"
        ws[f"M{row}"] = "No"
        ws[f"N{row}"] = "No"
        ws[f"O{row}"] = ""
        ws[f"P{row}"] = "No"
        ws[f"Q{row}"] = data.get("Notes", "")
    wb.save(path)

def update_row_in_excel(wb, ws, path, biz, **kwargs):
    row_num = biz["_row"]
    col_count = ws.max_column
    
    if col_count >= 14:
        col_map = {"Msg Sent?": "I", "Reply Came?": "J", "Follow Up Sent?": "K", "Follow Up Date": "L", "Converted?": "M", "Notes": "N"}
    else:
        col_map = {"Msg Sent?": "L", "Reply Came?": "M", "Follow Up Sent?": "N", "Follow Up Date": "O", "Converted?": "P", "Notes": "Q"}
    
    for key, val in kwargs.items():
        if key in col_map:
            ws[f"{col_map[key]}{row_num}"] = val
    wb.save(path)

def check_duplicate_in_excel(businesses, name, phone):
    for b in businesses:
        if name and str(b.get("Business Name", "")).lower() == name.lower():
            return b
        if phone and str(b.get("Phone No", "")) == phone:
            return b
    return None
